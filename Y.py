import os

# Check if openpyxl and xlrd are installed
try:
    from openpyxl import load_workbook
    import xlrd
    from openpyxl import Workbook
except ImportError:
    print("Please install openpyxl and xlrd using: pip install openpyxl xlrd")
    exit()

# Path to the Excel file
file_path = "your_file.xlsx"  # Change this to your file

# Set to store unique column names
unique_columns = set()

# Function to read .xlsx files using openpyxl
def read_xlsx(file_path):
    try:
        wb = load_workbook(file_path, data_only=True)  # Load workbook
        sheet = wb.active  # Get the first sheet

        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):  # Ignore empty or non-string values
                    unique_columns.add(cell.strip())

        print(f"Processed {file_path} successfully.")
    except Exception as e:
        print(f"Error reading {file_path}: {e}")

# Function to read .xls files using xlrd
def read_xls(file_path):
    try:
        wb = xlrd.open_workbook(file_path)
        sheet = wb.sheet_by_index(0)  # Get the first sheet

        for row_idx in range(sheet.nrows):
            for cell in sheet.row_values(row_idx):
                if cell and isinstance(cell, str):  # Ignore empty or non-string values
                    unique_columns.add(cell.strip())

        print(f"Processed {file_path} successfully.")
    except Exception as e:
        print(f"Error reading {file_path}: {e}")

# Check file extension and process accordingly
if file_path.endswith(".xlsx"):
    read_xlsx(file_path)
elif file_path.endswith(".xls"):
    read_xls(file_path)
else:
    print("Unsupported file format. Please provide a .xlsx or .xls file.")
    exit()

# Save unique column names to a new Excel file
output_file = "master_ciq.xlsx"
wb = Workbook()
ws = wb.active
ws.append(["Master CIQ Columns"])  # Add header

for col in sorted(unique_columns):  # Sort for consistency
    ws.append([col])

wb.save(output_file)

print(f"Master CIQ saved to: {output_file}")
